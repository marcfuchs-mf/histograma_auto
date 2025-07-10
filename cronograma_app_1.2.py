import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import unicodedata
import re
import math
from io import BytesIO

# Configuração da página
st.set_page_config(page_title="Gerador de Cronograma MF", layout="wide")
st.title("\U0001F4CA Gerador de Cronograma por Disciplina - MF")

# Entradas principais
empresa = st.text_input("Nome da empresa")
industria = st.selectbox("Tipo de indústria", [
    "Farmacêutica", "Alimentícia", "Automotiva", "Máquinas e Equipamentos",
    "Química", "Papel e Celulose", "Hospitais", "Aeroportos", "Mineradoras", "Outras"
])
tipo_projeto = st.radio("Tipo de Projeto", ["Greenfield", "Brownfield"])
fator_projeto = 1.5 if tipo_projeto == "Greenfield" else 1.0

m2 = st.number_input("Área do projeto (em m²)", min_value=0, step=1)
qtd_meses = st.slider("Quantidade de meses no cronograma", 1, 20, 12)

# Parâmetros de custo hora
st.markdown("### \U0001F4B2 Parâmetros de custo hora")
preco_coordenacao = st.number_input("Preço Hora da Coordenação (R$)", min_value=0.0, value=375.0)
valor_junior = st.number_input("Valor hora do profissional **Júnior** (R$)", min_value=0.0, value=275.0)
valor_senior = st.number_input("Valor hora do profissional **Sênior** (R$)", min_value=0.0, value=315.0)

# Multiplicadores por tipo de indústria
multiplicadores_industria = {
    "Farmacêutica": 1.5, "Alimentícia": 1.5, "Automotiva": 1.5, "Máquinas e Equipamentos": 1.5,
    "Química": 1.5, "Papel e Celulose": 1.5, "Hospitais": 1.5, "Aeroportos": 1.5,
    "Mineradoras": 1.5, "Outras": 1.2
}
horas_estimadas = round(m2 * multiplicadores_industria.get(industria, 1.2) * fator_projeto, 2)

# Disciplinas
disciplinas_percentuais = {
    "Coordenação": (0, 100),
    "Arquitetura": (0, 100),
    "Planejamento": (0, 100),
    "Infraestrutura": (10, 40),
    "Concreto": (15, 30),
    "Metálica": (20, 45),
    "Hidráulica": (30, 60),
    "HVAC": (50, 70),
    "BIM": (70, 100),
    "Elétrica": (40, 70),
    "Tubulação": (40, 80),
    "Automação": (60, 90),
    "Incêndio": (20, 50)
}

st.subheader("\U0001F4CC Selecione as disciplinas e preencha os dados:")
disciplinas_escolhidas = []

with st.form("formulario_disciplinas"):
    for nome, (p_ini, p_fim) in disciplinas_percentuais.items():
        col1, col2, col3, col4, col5 = st.columns([3, 2, 2, 2, 2])
        with col1:
            incluir = st.checkbox(nome, key=f"check_{nome}")
        if incluir:
            mes_ini = max(1, math.floor(qtd_meses * p_ini / 100))
            mes_fim = min(qtd_meses, math.ceil(qtd_meses * p_fim / 100))
            with col2:
                inicio = st.number_input(f"Início '{nome}'", 1, qtd_meses, mes_ini, key=f"ini_{nome}")
            with col3:
                fim = st.number_input(f"Término '{nome}'", 1, qtd_meses, mes_fim, key=f"fim_{nome}")
            if nome == "Coordenação":
                n_junior = 0
                with col5:
                    n_senior = st.number_input(f"Sênior '{nome}'", 0, 100, 0, key=f"sen_{nome}")
            else:
                with col4:
                    n_junior = st.number_input(f"Júnior '{nome}'", 0, 100, 0, key=f"jun_{nome}")
                with col5:
                    n_senior = st.number_input(f"Sênior '{nome}'", 0, 100, 0, key=f"sen_{nome}")
            disciplinas_escolhidas.append((nome, inicio, fim, n_junior, n_senior))
    gerar = st.form_submit_button("\U0001F4E5 Gerar Planilha")

def limpar_nome(nome):
    nome = unicodedata.normalize('NFKD', nome).encode('ASCII', 'ignore').decode('utf-8')
    nome = re.sub(r'\W+', '_', nome.lower())
    return nome.strip('_')

if gerar and empresa:
    disciplinas_invalidas = [
        nome for nome, _, _, n_j, n_s in disciplinas_escolhidas
        if n_j + n_s == 0 and nome != "Coordenação"
    ]

    if disciplinas_invalidas:
        st.error(f"❌ As seguintes disciplinas precisam de pelo menos 1 profissional: {', '.join(disciplinas_invalidas)}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cronograma Disciplinas"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=qtd_meses + 4)
        ws.cell(row=1, column=1, value=f"Cronograma - {empresa} - Indústria {industria}").font = Font(bold=True, size=14)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=qtd_meses + 4)
        ws.cell(row=2, column=1, value=f"Área total do projeto: {m2} m² | Tipo: {tipo_projeto}").font = Font(italic=True)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=qtd_meses + 4)
        ws.cell(row=3, column=1, value=f"Horas estimadas do projeto: {horas_estimadas:.0f}").font = Font(italic=True)

        ws.cell(row=4, column=1, value="Disciplina/Mês").font = Font(bold=True)
        for mes in range(1, qtd_meses + 1):
            ws.cell(row=4, column=mes + 1, value=f"Mês {mes}").font = Font(bold=True)
        ws.cell(row=4, column=qtd_meses + 2, value="Preço Hora Mix (R$)").font = Font(bold=True)
        ws.cell(row=4, column=qtd_meses + 3, value="Custo Total (R$)").font = Font(bold=True)

        linha = 5
        total_geral_custo = 0
        total_profissionais = sum((n_junior + n_senior if nome != "Coordenação" else n_senior) * (fim - ini + 1) for nome, ini, fim, n_junior, n_senior in disciplinas_escolhidas)
        horas_por_profissional_mes = horas_estimadas / total_profissionais if total_profissionais > 0 else 0

        col_somas = [0] * qtd_meses

        for nome, ini, fim, n_junior, n_senior in disciplinas_escolhidas:
            total_meses = fim - ini + 1
            total_profissionais_disc = (n_junior + n_senior if nome != "Coordenação" else n_senior)
            total_horas_disciplina = total_meses * total_profissionais_disc * horas_por_profissional_mes

            preco_hora_mix = preco_coordenacao if nome == "Coordenação" else round((n_junior * valor_junior + n_senior * valor_senior) / (n_junior + n_senior), 2) if (n_junior + n_senior) > 0 else 0
            custo_total_disciplina = round(preco_hora_mix * total_horas_disciplina, 2)
            total_geral_custo += custo_total_disciplina

            cell_disc = ws.cell(row=linha, column=1, value=nome)
            cell_disc.font = Font(bold=True)
            for i, col in enumerate(range(ini + 1, fim + 2)):
                valor = round(total_horas_disciplina / total_meses)
                ws.cell(row=linha, column=col, value=int(valor))
                col_somas[col - 2] += valor
            ws.cell(row=linha, column=qtd_meses + 2, value=preco_hora_mix)
            ws.cell(row=linha, column=qtd_meses + 3, value=custo_total_disciplina)

            def format_prof_linha(texto, valor_unit, valor_total):
                ws.cell(row=linha, column=1, value=texto)
                for col in range(ini + 1, fim + 2):
                    ws.cell(row=linha, column=col, value=int(round(horas_por_profissional_mes * (n_senior if 'Sênior' in texto else n_junior))))
                ws.cell(row=linha, column=qtd_meses + 2, value=valor_unit)
                ws.cell(row=linha, column=qtd_meses + 3, value=valor_total)

            if nome == "Coordenação":
                linha += 1
                format_prof_linha("Horas Sênior", preco_coordenacao, round(preco_coordenacao * horas_por_profissional_mes * n_senior * total_meses, 2))
            else:
                linha += 1
                format_prof_linha("Horas Júnior", valor_junior, round(valor_junior * horas_por_profissional_mes * n_junior * total_meses, 2))
                linha += 1
                format_prof_linha("Horas Sênior", valor_senior, round(valor_senior * horas_por_profissional_mes * n_senior * total_meses, 2))

            linha += 1

        ws.cell(row=linha, column=1, value="Total por mês").font = Font(bold=True)
        for i, total in enumerate(col_somas):
            ws.cell(row=linha, column=i + 2, value=int(round(total)))

        ws.cell(row=1, column=qtd_meses + 3, value=f"Valor Total: R$ {total_geral_custo:,.2f}").font = Font(bold=True)

        for i, col in enumerate(ws.columns, start=1):
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 2

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        st.success("\u2705 Planilha gerada com sucesso!")
        st.download_button(
            label="\U0001F4E4 Baixar Excel",
            data=buffer,
            file_name=f"cronograma_{limpar_nome(empresa)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

st.markdown("---")
st.markdown("### \U0001F6D1 Como encerrar o Streamlit")
st.code("CTRL + C", language="bash")

# Rodar no terminal:
# python -m streamlit run cronograma_app_1.2.py
